from ExcelSearch import rep, REPLACE
import os
import openpyxl
import re
from colorama import Fore, Style

def red_text(text):
    return Fore.RED+Style.BRIGHT+text+Style.RESET_ALL

def blue_text(text):
    return Fore.BLUE+Style.BRIGHT+text+Style.RESET_ALL

pattern = r'=SUM\(\w\d+:(\w\d+)\)'

def checkSum(path):
    for root, dirs, files in os.walk(path):
        for file in files:
            if not file.endswith(".xlsx"):
                continue
            full_path = os.path.join(root, file)
            full_path = full_path.replace('\\', '/')
            try:
                wb = openpyxl.load_workbook(full_path)
            except PermissionError as e:
                print(e)
                continue
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        ret = re.match(pattern, str(cell.value))
                        if ret != None:
                            start = int(ret.group(1)[1:])
                            end = int(cell.coordinate[1:])
                            if end - start != 1:
                                print(f'在 "{blue_text(full_path)}" 中的 "{red_text(cell.coordinate)}" 的求和 "{cell.value}" 可能存在问题')
                                rep_path = f'file:///{rep(full_path, REPLACE)}'
                                if ' ' in rep_path:
                                    rep_path = '文件名包含特殊字符，无法生成链接'
                                # print(f'ctrl+左键点此打开文件：{rep_path}\n')

if __name__ == '__main__':
    checkSum('D:/')
