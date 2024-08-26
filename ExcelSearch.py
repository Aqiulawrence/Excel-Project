import os
from openpyxl import load_workbook
from tkinter import messagebox

REPLACE = {
    '%20': '+',
    ' ': '%20',
    '!': '%21',
    "'": '%27',
    '(': '%28',
    ')': '%29',
    '~': '%7E',
    '%00': r'\x00'
}

def rep(rawstr, dict_rep):
    for i in dict_rep:
        rawstr = rawstr.replace(i, dict_rep[i])
    return rawstr

def find_data_in_single_excel(folder_path, data_to_find):
    try:
        file_path = folder_path
        if file_path.endswith('.xlsx'):
            workbook = load_workbook(file_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if data_to_find.lower() in str(cell.value).lower():
                            print(f'在 "{file_path}" 中的 "{cell.coordinate}" 找到 "{data_to_find}"')
                            print('当前行的内容：', end='')
                            for c in row:
                                if c.value != None:
                                    print(f"{c.coordinate[0]}:'{c.value}'", end=', ')
                            print('\n')
    except PermissionError:
        messagebox.showerror('错误', f'{folder_path} 拒绝访问。')
        return False
    return True

def find_data_in_multiple_excel(folder_path, data_to_find):
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if not file.endswith('.xlsx'):
                continue
            full_path = os.path.join(root, file)
            full_path = full_path.replace('\\', '/')
            try:
                workbook = load_workbook(full_path)
            except PermissionError as e:
                print(e)
                continue
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if data_to_find.lower() in str(cell.value).lower():
                            rep_path = f'file:///{rep(full_path, REPLACE)}'
                            if ' ' in rep_path:
                                rep_path = '文件名包含特殊字符，无法生成链接'
                            print(f'在 "{full_path}" 中的 "{cell.coordinate}" 找到 "{data_to_find}"，ctrl+左键点此打开文件：{rep_path}')
                            print('当前行的内容：', end='')
                            for c in row:
                                if c.value != None:
                                    print(f"{c.coordinate[0]}:'{c.value}'", end=', ')
                            print('\n')

if __name__ == '__main__':
    folder_path = r'D:/'
    data_to_find = '111'
    find_data_in_multiple_excel(folder_path, data_to_find)
    # find_data_in_single_excel(r'D:\Excel Project\excel\test.xlsx', data_to_find)