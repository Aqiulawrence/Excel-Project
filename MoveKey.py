import openpyxl
import re
from tkinter import messagebox

def main(file, start, end, move_to):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    pattern = r'\s[A-Za-z]+\s|\s[A-Za-z]+|[A-Za-z]+\s'
    now = start[0]
    target = move_to[0]
    target_y = int(move_to[1:])
    start = int(start[1:])
    end = int(end[1:])
    count = 0
    for i in range(start, end+1):
        cell = ws[f'{now}{i}']
        cell_value = cell.value.split('/')
        if len(cell_value) == 1:
            num = re.findall(r'\d+', cell_value[0])
            if num:
                if len(num) == 1:
                    ws[f'{target}{target_y+count}'] = ''.join(num)
                else:
                    num = re.sub(pattern, '', cell_value[0])
                    ws[f'{target}{target_y+count}'] = num
        else:
            ws[f'{target}{target_y+count}'] = cell_value[1]
        count += 1

    try: wb.save(file)
    except:
        messagebox.showerror('错误', '插入失败！请检查待插入的Excel文件是否已关闭！')
        raise

if __name__ == '__main__':
    main('test.xlsx', 'B1', 'B8', 'E9')